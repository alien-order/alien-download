from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os


class QueryGenerator:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("쿼리 생성기")
        self.root.geometry("1000x700")
        
        self.setup_styles()
        self.setup_ui()
        
        self.excel_data = []
        
    def setup_styles(self):
        """UI 스타일 설정"""
        self.root.configure(bg='#f0f0f0')
        
        style = ttk.Style()
        style.theme_use('clam')
        
        # 버튼 스타일
        style.configure('Action.TButton', 
                       font=('Arial', 10, 'bold'),
                       padding=(10, 5))
    
    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill="both", expand=True)
        
        # 설정 프레임
        config_frame = ttk.LabelFrame(main_frame, text="📋 설정", padding="15")
        config_frame.pack(fill="x", pady=(0, 15))
        
        # 엑셀 파일 선택
        file_row = ttk.Frame(config_frame)
        file_row.pack(fill="x", pady=5)
        ttk.Label(file_row, text="엑셀 파일:", font=('Arial', 9, 'bold')).pack(side="left", padx=(0, 10))
        self.file_path_var = tk.StringVar()
        file_entry = ttk.Entry(file_row, textvariable=self.file_path_var, font=('Arial', 9))
        file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        ttk.Button(file_row, text="📂 찾기", command=self.select_excel_file, 
                  style='Action.TButton').pack(side="right")
        
        # 시작 셀 설정
        cell_row = ttk.Frame(config_frame)
        cell_row.pack(fill="x", pady=5)
        ttk.Label(cell_row, text="시작 셀:", font=('Arial', 9, 'bold')).pack(side="left", padx=(0, 10))
        self.start_cell_var = tk.StringVar(value="A2")
        ttk.Entry(cell_row, textvariable=self.start_cell_var, width=10, font=('Arial', 9)).pack(side="left", padx=(0, 20))
        
        ttk.Label(cell_row, text="배치 크기:", font=('Arial', 9, 'bold')).pack(side="left", padx=(0, 10))
        self.batch_size_var = tk.StringVar(value="1000")
        ttk.Entry(cell_row, textvariable=self.batch_size_var, width=8, font=('Arial', 9)).pack(side="left", padx=(0, 10))
        
        ttk.Button(cell_row, text="📖 엑셀 읽기", command=self.load_excel_data,
                  style='Action.TButton').pack(side="right", padx=(20, 0))
        
        # 쿼리 템플릿 프레임
        template_frame = ttk.LabelFrame(main_frame, text="📝 쿼리 템플릿", padding="15")
        template_frame.pack(fill="x", pady=(0, 15))
        
        ttk.Label(template_frame, text="{VALUES} 부분에 값들이 삽입됩니다",
                 font=('Arial', 9, 'bold')).pack(anchor=tk.W, pady=(0, 5))
        
        template_container = ttk.Frame(template_frame)
        template_container.pack(fill="x")
        
        self.query_template = tk.Text(template_container, height=8, wrap=tk.WORD, 
                                     font=('Consolas', 10))
        template_scrollbar = ttk.Scrollbar(template_container, orient=tk.VERTICAL, 
                                          command=self.query_template.yview)
        self.query_template.configure(yscrollcommand=template_scrollbar.set)
        
        
        self.query_template.pack(side=tk.LEFT, fill="both", expand=True)
        template_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 제어 버튼
        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill="x", pady=(0, 15))
        
        ttk.Button(control_frame, text="🔧 쿼리 생성", command=self.generate_query,
                  style='Action.TButton').pack(side="left", padx=(0, 10))
        ttk.Button(control_frame, text="📋 복사", command=self.copy_query,
                  style='Action.TButton').pack(side="left", padx=(0, 10))
        ttk.Button(control_frame, text="💾 저장", command=self.save_query,
                  style='Action.TButton').pack(side="left")
        
        # 상태 표시
        self.status_var = tk.StringVar(value="엑셀 파일을 선택하고 읽기 버튼을 눌러주세요")
        ttk.Label(main_frame, textvariable=self.status_var, 
                 font=('Arial', 9), foreground="#666666").pack(pady=(0, 10))
        
        # 생성된 쿼리 출력 영역
        result_frame = ttk.LabelFrame(main_frame, text="📄 생성된 쿼리", padding="15")
        result_frame.pack(fill="both", expand=True)
        
        result_container = ttk.Frame(result_frame)
        result_container.pack(fill="both", expand=True)
        
        self.result_text = tk.Text(result_container, wrap=tk.WORD, 
                                  font=('Consolas', 9), bg='#f8f8f8')
        result_scrollbar = ttk.Scrollbar(result_container, orient=tk.VERTICAL, 
                                        command=self.result_text.yview)
        self.result_text.configure(yscrollcommand=result_scrollbar.set)
        
        self.result_text.pack(side=tk.LEFT, fill="both", expand=True)
        result_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    def select_excel_file(self):
        file_path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.status_var.set(f"파일 선택됨: {os.path.basename(file_path)}")
    
    def parse_cell_reference(self, cell_ref):
        """A1 형식의 셀 참조를 행과 열 인덱스로 변환"""
        cell_ref = cell_ref.upper().strip()
        
        col_str = ""
        row_str = ""
        
        for char in cell_ref:
            if char.isalpha():
                col_str += char
            elif char.isdigit():
                row_str += char
        
        if not col_str or not row_str:
            raise ValueError("올바른 셀 참조가 아닙니다 (예: A1, B10)")
        
        # 컬럼을 숫자로 변환 (A=0, B=1, ...)
        col_index = 0
        for char in col_str:
            col_index = col_index * 26 + (ord(char) - ord('A') + 1)
        col_index -= 1  # 0-based
        
        # 행을 숫자로 변환 (1-based에서 0-based로)
        row_index = int(row_str) - 1
        
        return row_index, col_index
    
    def load_excel_data(self):
        """엑셀 파일에서 데이터 읽기"""
        if not self.file_path_var.get():
            messagebox.showerror("오류", "엑셀 파일을 선택해주세요.")
            return

        if not os.path.exists(self.file_path_var.get()):
            messagebox.showerror("오류", "선택한 엑셀 파일이 존재하지 않습니다.")
            return

        try:
            # 셀 참조 파싱
            start_row, start_col = self.parse_cell_reference(self.start_cell_var.get())

            # openpyxl로 엑셀 읽기 (read_only 모드로 메모리 효율적)
            wb = load_workbook(self.file_path_var.get(), read_only=True, data_only=True)
            ws = wb.active

            # 시작 셀부터 데이터 추출 (openpyxl은 1-based index)
            self.excel_data = []
            col_idx = start_col + 1  # openpyxl은 1부터 시작

            for row_idx, row in enumerate(ws.iter_rows(min_row=start_row + 1, min_col=col_idx, max_col=col_idx), start=start_row):
                cell_value = row[0].value
                if cell_value is not None:
                    self.excel_data.append(str(cell_value))

            wb.close()

            self.status_var.set(f"✅ {len(self.excel_data)}개 데이터 로드 완료")

            # 미리보기 표시
            preview = ", ".join(self.excel_data[:5])
            if len(self.excel_data) > 5:
                preview += f"... (총 {len(self.excel_data)}개)"

            self.result_text.delete(1.0, tk.END)
            self.result_text.insert(tk.END, f"데이터 미리보기:\n{preview}")

        except Exception as e:
            messagebox.showerror("오류", f"엑셀 파일 읽기 실패: {str(e)}")
            self.status_var.set("❌ 파일 읽기 실패")
    
    def generate_query(self):
        """쿼리 생성"""
        if not self.excel_data:
            messagebox.showerror("오류", "먼저 엑셀 데이터를 로드해주세요.")
            return
        
        try:
            batch_size = int(self.batch_size_var.get())
            if batch_size <= 0:
                raise ValueError("배치 크기는 1 이상이어야 합니다")
        except ValueError as e:
            messagebox.showerror("오류", f"올바른 배치 크기를 입력해주세요: {str(e)}")
            return
        
        template = self.query_template.get(1.0, tk.END).strip()
        if not template or "{VALUES}" not in template:
            messagebox.showerror("오류", "쿼리 템플릿에 {VALUES} 플레이스홀더가 필요합니다.")
            return
        
        try:
            self.status_var.set("🔧 쿼리 생성 중...")
            self.root.update()
            
            # 데이터 개수 확인
            data_count = len(self.excel_data)
            
            if data_count <= batch_size:
                # 배치 크기 이하면 단일 쿼리
                values_str = ", ".join([f"'{value}'" for value in self.excel_data])
                final_query = template.replace("{VALUES}", values_str)
                
                # 결과 표시
                self.result_text.delete(1.0, tk.END)
                self.result_text.insert(tk.END, final_query)
                
                # 통계 표시
                stats = f"✅ 단일 쿼리 생성 완료 - {data_count}개 데이터"
                self.status_var.set(stats)
                
            else:
                # 배치 크기 초과시 UNION ALL 사용
                batches = []
                for i in range(0, data_count, batch_size):
                    batch = self.excel_data[i:i+batch_size]
                    # 값들을 문자열로 포맷팅 (작은따옴표 추가)
                    values_str = ", ".join([f"'{value}'" for value in batch])
                    batches.append(values_str)
                
                # UNION ALL로 연결된 쿼리 생성
                queries = []
                for batch_values in batches:
                    query = template.replace("{VALUES}", batch_values)
                    queries.append(query)
                
                final_query = "\n\nUNION ALL\n\n".join(queries)
                
                # 결과 표시
                self.result_text.delete(1.0, tk.END)
                self.result_text.insert(tk.END, final_query)
                
                # 통계 표시
                stats = f"✅ UNION ALL 쿼리 생성 완료 - {data_count}개 데이터, {len(batches)}개 배치"
                self.status_var.set(stats)
            
        except Exception as e:
            messagebox.showerror("오류", f"쿼리 생성 실패: {str(e)}")
            self.status_var.set("❌ 쿼리 생성 실패")
    
    def copy_query(self):
        """생성된 쿼리를 클립보드에 복사"""
        query = self.result_text.get(1.0, tk.END).strip()
        
        if not query or query == "데이터 미리보기:":
            messagebox.showwarning("경고", "복사할 쿼리가 없습니다. 먼저 쿼리를 생성해주세요.")
            return
        
        try:
            # tkinter 내장 클립보드 사용
            self.root.clipboard_clear()
            self.root.clipboard_append(query)
            self.root.update()  # 클립보드 업데이트 보장
            self.status_var.set("클립보드에 복사됨")
            messagebox.showinfo("성공", "쿼리가 클립보드에 복사되었습니다!")
        except Exception as e:
            messagebox.showerror("오류", f"클립보드 복사 실패: {str(e)}")
    
    def save_query(self):
        """생성된 쿼리를 파일로 저장"""
        query = self.result_text.get(1.0, tk.END).strip()
        
        if not query or query == "데이터 미리보기:":
            messagebox.showwarning("경고", "저장할 쿼리가 없습니다. 먼저 쿼리를 생성해주세요.")
            return
        
        try:
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"generated_query_{timestamp}.sql"
            
            file_path = filedialog.asksaveasfilename(
                title="쿼리 저장",
                defaultextension=".sql",
                initialname=default_filename,
                filetypes=[("SQL files", "*.sql"), ("Text files", "*.txt"), ("All files", "*.*")]
            )
            
            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(query)
                self.status_var.set(f"💾 파일 저장됨: {os.path.basename(file_path)}")
                messagebox.showinfo("성공", f"쿼리가 저장되었습니다:\n{file_path}")
                
        except Exception as e:
            messagebox.showerror("오류", f"파일 저장 실패: {str(e)}")
    
    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = QueryGenerator()
    app.run()